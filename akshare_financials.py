#!/usr/bin/env python3
"""Fetch key financial metrics for A-share companies via AKShare.

默认展示 2024 年年报至 2025 年三季报的指标，可按需指定股票和期间，
支持导出 Excel 并包含实时估值信息。
"""

from __future__ import annotations

import argparse
from dataclasses import dataclass
from datetime import datetime
import time
from pathlib import Path
from typing import Dict, Iterable, Optional, Tuple

import akshare as ak
import pandas as pd
from pandas.tseries.offsets import DateOffset
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# 目标报告期，可按需修改或通过 CLI 传入
DEFAULT_REPORT_DATES: Dict[str, str] = {
    "2025三季报": "2025-09-30",
    "2025中报": "2025-06-30",
    "2025一季报": "2025-03-31",
    "2024年报": "2024-12-31",
}

DISPLAY_METRIC_ORDER = [
    # 利润表
    "营业收入",
    "营业收入增长率",
    "毛利率",
    "营业利润率",
    "净利润率",
    "净利润",
    "归母净利润",
    # ROE
    "归母净利润率",
    "总资产周转率",
    "权益乘数",
    "ROE",
    # 现金流量表
    "经营活动现金流量净额",
    "净利润现金保障倍数",
    "投资活动现金流量净额",
    "销售商品提供劳务收到的现金",
    "现金收入比率",
    "资本性支出",
    "自由现金流",
    # 资产负债表
    "资产负债率",
    "流动比率",
    "净资产",
    # 估值
    "市值",
    "PE",
    "PB",
    "PS",
    "估值",
    "估值时间",
    # 最后估值
    "最新PE",
    "最新PB",
    "最新PS",
    "最新估值",
    "最新估值时间",
]

PERCENT_METRICS = {
    "营业收入增长率",
    "毛利率",
    "营业利润率",
    "净利润率",
    "归母净利润率",
    "现金收入比率",
    "资产负债率",
    "ROE",
}

MULTIPLIER_METRICS = {"PE", "PB", "PS", "最新PE", "最新PB", "最新PS"}

CURRENCY_KEYWORDS = (
    "收入",
    "现金",
    "利润",
    "净额",
    "净流量",
    "支出",
    "资产",
    "资本",
    "权益",
    "成本",
    "费用",
    "市值",
)

LATEST_ONLY_METRICS = {
    "最新PE": ("2025三季报", "本期"),
    "最新PB": ("2025三季报", "本期"),
    "最新PS": ("2025三季报", "本期"),
    "最新估值": ("2025三季报", "本期"),
    "最新估值时间": ("2025三季报", "本期"),
}

SECTION_DEFINITIONS = [
    (
        "利润表",
        [
            "营业收入",
            "营业收入增长率",
            "毛利率",
            "营业利润率",
            "净利润率",
            "净利润",
            "归母净利润",
        ],
    ),
    (
        "ROE",
        [
            "归母净利润率",
            "总资产周转率",
            "权益乘数",
            "ROE",
        ],
    ),
    (
        "现金流量表",
        [
            "经营活动现金流量净额",
            "净利润现金保障倍数",
            "投资活动现金流量净额",
            "销售商品提供劳务收到的现金",
            "现金收入比率",
            "资本性支出",
            "自由现金流",
        ],
    ),
    (
        "资产负债表",
        [
            "资产负债率",
            "流动比率",
            "净资产",
        ],
    ),
    (
        "估值",
        [
            "市值",
            "PE",
            "PB",
            "PS",
            "估值",
            "估值时间",
        ],
    ),
    (
        "最后估值",
        [
            "最新PE",
            "最新PB",
            "最新PS",
            "最新估值",
            "最新估值时间",
        ],
    ),
]

INVALID_FILENAME_CHARS = set('\\/:*?"<>|')


@dataclass(frozen=True)
class LatestValuation:
    pe: Optional[float]
    pb: Optional[float]
    ps: Optional[float]
    timestamp: Optional[str]
    price: Optional[float]
    market_cap: Optional[float]


@dataclass(frozen=True)
class Dataset:
    symbol: str
    stock_code: str
    company_name: str
    profit: pd.DataFrame
    cashflow: pd.DataFrame
    balance: pd.DataFrame
    latest_price: Optional[float]
    total_shares: Optional[float]
    total_market_cap: Optional[float]
    latest_valuation: Optional[LatestValuation]


def convert_numeric(df: pd.DataFrame) -> pd.DataFrame:
    for column in df.columns:
        if df[column].dtype == object:
            converted = pd.to_numeric(df[column], errors="coerce")
            if converted.notna().any():
                df[column] = converted
    return df


def sanitize_filename_component(text: str) -> str:
    sanitized = ''.join(ch for ch in text if ch not in INVALID_FILENAME_CHARS)
    return sanitized.strip()


def strip_exchange_prefix(code: str) -> str:
    raw = code.strip().lower()
    if raw.startswith(("sh", "sz", "bj")) and len(raw) > 2:
        return raw[2:]
    return raw


def default_excel_filename(dataset: Dataset) -> str:
    safe_company = sanitize_filename_component(dataset.company_name) or strip_exchange_prefix(dataset.stock_code).upper()
    code_tag = strip_exchange_prefix(dataset.stock_code).upper()
    return f"{safe_company}_{code_tag}.xlsx"


def fetch_company_overview(base_symbol: str) -> Tuple[Optional[str], Optional[float], Optional[float], Optional[float]]:
    try:
        info_df = ak.stock_individual_info_em(symbol=base_symbol)
    except Exception:
        return None, None, None, None
    if info_df is None or info_df.empty:
        return None, None, None, None

    info_map = {
        str(row["item"]).strip(): row["value"]
        for _, row in info_df.iterrows()
    }

    def to_float(value: object) -> Optional[float]:
        if value is None or value == "":
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    name = info_map.get("股票简称") or info_map.get("证券简称")
    company_name = str(name).strip() if name else None
    latest_price = to_float(info_map.get("最新"))
    total_shares = to_float(info_map.get("总股本"))
    total_market_cap = to_float(info_map.get("总市值"))

    return company_name, latest_price, total_shares, total_market_cap


def fetch_latest_valuation(base_symbol: str, retries: int = 3, delay: float = 1.5) -> Optional[LatestValuation]:
    last_error: Optional[Exception] = None
    for attempt in range(1, retries + 1):
        try:
            spot_df = ak.stock_zh_a_spot_em()
        except Exception as err:
            last_error = err
            print(f"警告: 获取最新估值失败（第 {attempt} 次），原因: {err}")
            if attempt < retries:
                time.sleep(delay)
            continue

        if spot_df is None or spot_df.empty:
            print(f"警告: 获取最新估值失败（第 {attempt} 次），返回数据为空")
            if attempt < retries:
                time.sleep(delay)
            continue

        row = spot_df.loc[spot_df["代码"] == base_symbol]
        if row.empty:
            print(f"警告: 最新估值数据中未找到代码 {base_symbol}")
            return None

        record = row.iloc[0]

        def to_float(value: object) -> Optional[float]:
            if value is None or value == "":
                return None
            try:
                return float(value)
            except (TypeError, ValueError):
                return None

        pe = to_float(record.get("市盈率-动态")) or to_float(record.get("市盈率(TTM)"))
        pb = to_float(record.get("市净率"))
        ps = to_float(record.get("市销率")) or to_float(record.get("市销率(TTM)"))
        price = to_float(record.get("最新价")) or to_float(record.get("现价"))
        market_cap = (
            to_float(record.get("总市值"))
            or to_float(record.get("总市值-实时"))
            or to_float(record.get("总市值-最新"))
        )

        timestamp = record.get("更新时间") or record.get("数据日期")
        if timestamp is None or str(timestamp).strip() == "":
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        else:
            timestamp = str(timestamp)

        return LatestValuation(pe=pe, pb=pb, ps=ps, timestamp=timestamp, price=price, market_cap=market_cap)

    if last_error is not None:
        print("警告: 多次尝试获取最新估值失败，已放弃。")
    return None


def is_currency_metric(metric: str) -> bool:
    base = metric.replace("-去年同期", "")
    if base in PERCENT_METRICS:
        return False
    if base.endswith("率") or "比率" in base:
        return False
    if base.endswith("倍数") or "倍数" in base:
        return False
    return any(keyword in base for keyword in CURRENCY_KEYWORDS)


def format_value(metric: str, value: Optional[float]) -> str:
    if value is None:
        return "-"
    if isinstance(value, str):
        text = value.strip()
        return text if text else "-"
    if isinstance(value, (datetime, pd.Timestamp)):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if pd.isna(value):
        return "-"
    base_metric = metric.replace("-去年同期", "")
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return str(value)
    if base_metric in PERCENT_METRICS:
        return f"{numeric * 100:.2f}%"
    if base_metric in MULTIPLIER_METRICS:
        return f"{numeric:.2f}倍"
    if base_metric.endswith("倍数") or "倍数" in base_metric:
        return f"{numeric:.2f}倍"
    if "乘数" in base_metric:
        return f"{numeric:.2f}倍"
    if is_currency_metric(metric):
        absolute = abs(numeric)
        if absolute >= 1e8:
            return f"{numeric / 1e8:.2f}亿"
        if absolute >= 1e4:
            return f"{numeric / 1e4:.2f}万"
        return f"{numeric:.2f}元"
    return f"{numeric:.2f}"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="获取上市公司主要财务指标（含同比）")
    parser.add_argument(
        "--symbol",
        default=None,
        help="单只股票代码，例：600418（沪）或 000XXX（深）(可与 --symbols 同用)",
    )
    parser.add_argument(
        "--symbols",
        nargs="+",
        default=None,
        help="股票代码列表，例：600418 000333 601012",
    )
    parser.add_argument(
        "--excel",
        type=Path,
        default=None,
        help="将结果导出为 Excel 文件的路径（可选）",
    )
    parser.add_argument(
        "--report",
        nargs="*",
        default=None,
        metavar=("名称", "日期"),
        help=(
            "自定义报告期，例如 --report 2025Q3 2025-09-30 2025Q2 2025-06-30 。"
            "没有提供时使用默认的四个报告期。"
        ),
    )
    return parser.parse_args()


def build_report_dates(args_report: Optional[Iterable[str]]) -> Dict[str, str]:
    if not args_report:
        return DEFAULT_REPORT_DATES
    values = list(args_report)
    if len(values) % 2 != 0:
        raise ValueError("--report 参数必须成对出现：<名称> <日期>")
    iterator = iter(values)
    custom: Dict[str, str] = {}
    for name, date_str in zip(iterator, iterator):
        custom[name] = date_str
    return custom


def normalize_stock(code: str) -> str:
    raw = code.strip().lower()
    if raw.startswith(("sh", "sz", "bj")):
        return raw
    if len(raw) == 6 and raw.isdigit():
        prefix = "sh" if raw.startswith("6") else "sz" if raw[0] in {"0", "3"} else None
        if prefix:
            return f"{prefix}{raw}"
        if raw[0] in {"4", "8"}:
            return f"bj{raw}"
    raise ValueError(f"无法识别的股票代码格式：{code}")


def find_column(df: pd.DataFrame, candidates: Iterable[str]) -> Optional[str]:
    for name in candidates:
        if name in df.columns:
            return name
    return None


def get_value_at(df: pd.DataFrame, date_str: str, column: Optional[str]) -> Optional[float]:
    if column is None or column not in df.columns:
        return None
    ts = pd.Timestamp(date_str)
    if ts not in df.index:
        return None
    value = df.at[ts, column]
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def get_value(df: pd.DataFrame, report_date: str, column: Optional[str], years_offset: int = 0) -> Optional[float]:
    if column is None or column not in df.columns:
        return None
    idx = pd.Timestamp(report_date) - DateOffset(years=years_offset)
    if idx not in df.index:
        return None
    value = df.at[idx, column]
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def compute_period_market_caps(
    dataset: Dataset, report_dates: Dict[str, str]
) -> Dict[str, Dict[str, Optional[float]]]:
    total_shares = dataset.total_shares
    if total_shares in (None, 0) or pd.isna(total_shares):
        return {}

    if not report_dates:
        return {}

    report_ts = sorted(pd.Timestamp(date) for date in report_dates.values())
    start = (report_ts[0] - pd.DateOffset(days=370)).strftime("%Y%m%d")
    end = (report_ts[-1] + pd.DateOffset(days=5)).strftime("%Y%m%d")

    symbol = strip_exchange_prefix(dataset.stock_code)
    try:
        hist_df = ak.stock_zh_a_hist(symbol=symbol, start_date=start, end_date=end, adjust="")
    except Exception:
        return {}

    if hist_df is None or hist_df.empty or "收盘" not in hist_df.columns:
        return {}

    hist_df = hist_df.copy()
    hist_df["日期"] = pd.to_datetime(hist_df["日期"])
    hist_df = hist_df.set_index("日期").sort_index()
    close_series = hist_df["收盘"].astype(float)

    info: Dict[str, Dict[str, Optional[float]]] = {}
    for label, date_str in report_dates.items():
        ts = pd.Timestamp(date_str)
        relevant = close_series.loc[:ts]
        if relevant.empty:
            continue
        last_price = relevant.iloc[-1]
        price = float(last_price)
        price_time = relevant.index[-1]
        info[label] = {
            "market_cap": float(price) * float(total_shares),
            "price": price,
            "price_time": price_time,
        }

    return info


def safe_sub(minuend: Optional[float], subtrahend: Optional[float]) -> Optional[float]:
    if minuend is None or subtrahend is None:
        return None
    return minuend - subtrahend


def calc_growth(current: Optional[float], previous: Optional[float]) -> Optional[float]:
    if current is None or previous in (None, 0):
        return None
    try:
        return (current - previous) / abs(previous)
    except ZeroDivisionError:
        return None


def mean_valid(*values: Optional[float]) -> Optional[float]:
    valid = [v for v in values if v is not None]
    if not valid:
        return None
    return sum(valid) / len(valid)


def safe_product(values: Iterable[Optional[float]]) -> Optional[float]:
    result = 1.0
    for value in values:
        if value is None:
            return None
        result *= value
    return result


def calc_ttm(current: Optional[float], prev_same_period: Optional[float], prev_year_end: Optional[float]) -> Optional[float]:
    if current is None:
        return None
    if prev_same_period is None or prev_year_end is None:
        return current
    return current + (prev_year_end - prev_same_period)


def metric_base(metric: str) -> str:
    return metric.replace("-去年同期", "")


def get_section(metric: str) -> Optional[str]:
    base = metric_base(metric)
    for section_name, metrics in SECTION_DEFINITIONS:
        if base in metrics:
            return section_name
    return None


def insert_period_separators(df: pd.DataFrame, periods: Iterable[str]) -> Tuple[pd.DataFrame, list[str]]:
    result = df.copy()
    sep_columns: list[str] = []
    period_list = list(periods)
    if not period_list:
        return result, sep_columns

    for period_index, period in enumerate(period_list):
        if period_index == len(period_list) - 1:
            break

        period_columns = [
            col
            for col in result.columns
            if not str(col).startswith("__sep_")
            and col != "指标"
            and str(col).split("-", 1)[0] == period
        ]

        if not period_columns:
            continue

        last_col_name = period_columns[-1]
        last_position = result.columns.get_loc(last_col_name)
        sep_name = f"__sep_{len(sep_columns) + 1}"
        result.insert(last_position + 1, sep_name, [""] * len(result))
        sep_columns.append(sep_name)

    return result, sep_columns


def calc_display_length(text: str) -> int:
    if text is None:
        return 0
    return sum(2 if ord(ch) > 127 else 1 for ch in str(text))


def combine_summary(summary: pd.DataFrame, period_order: Iterable[str]) -> pd.DataFrame:
    if summary.empty:
        return pd.DataFrame()

    period_list = [period for period in period_order if period in summary.index]
    if not period_list:
        period_list = list(summary.index)

    metric_values: Dict[str, Dict[Tuple[str, str], Optional[float]]] = {}

    for period in period_list:
        row = summary.loc[period]
        for column, value in row.items():
            value_type = "去年同期" if column.endswith("-去年同期") else "本期"
            base_metric = metric_base(column)
            metric_values.setdefault(base_metric, {})[(period, value_type)] = value

    ordered_metrics: list[str] = []

    for _, metrics in SECTION_DEFINITIONS:
        for metric in metrics:
            if metric in metric_values and metric not in ordered_metrics:
                ordered_metrics.append(metric)

    for metric in DISPLAY_METRIC_ORDER:
        if metric in metric_values and metric not in ordered_metrics:
            ordered_metrics.append(metric)

    remaining_metrics = [metric for metric in metric_values if metric not in ordered_metrics]
    ordered_metrics.extend(sorted(remaining_metrics))

    columns: list[Tuple[str, str]] = []
    for period in period_list:
        for value_type in ("本期", "去年同期"):
            columns.append((period, value_type))

    columns = [column for column in columns if any(column in values for values in metric_values.values())]

    combined = pd.DataFrame(index=ordered_metrics, columns=pd.MultiIndex.from_tuples(columns, names=["期间", "口径"]))

    for metric in ordered_metrics:
        values = metric_values.get(metric, {})
        for column in columns:
            combined.at[metric, column] = values.get(column)

    def first_nonempty(metric: str) -> Optional[object]:
        if metric not in combined.index:
            return None
        for column in columns:
            value = combined.at[metric, column]
            if value not in (None, "") and not (isinstance(value, float) and pd.isna(value)):
                return value
        return None

    for metric, target in LATEST_ONLY_METRICS.items():
        if metric not in combined.index:
            continue

        period, col_type = target
        target_col = (period, col_type)
        if target_col not in combined.columns:
            combined[target_col] = None

        value = first_nonempty(metric)
        combined.at[metric, target_col] = value

        for column in columns:
            if column == target_col:
                continue
            combined.at[metric, column] = None

    combined = combined.dropna(axis=1, how="all")
    combined.index.name = "指标"
    return combined


def add_section_headers(table: pd.DataFrame) -> pd.DataFrame:
    if table.empty:
        return table

    new_index: list[str] = []
    new_data: Dict[str, list[object]] = {col: [] for col in table.columns}
    current_section: Optional[str] = None
    section_counter = 0

    for idx, row in table.iterrows():
        section = get_section(idx)
        if section and section != current_section:
            section_counter += 1
            header_label = f"【{section}】"
            new_index.append(header_label)
            for col in table.columns:
                new_data[col].append("")
            current_section = section

        new_index.append(idx)
        for col in table.columns:
            new_data[col].append(row[col])

    df = pd.DataFrame(new_data, index=new_index)
    df.index.name = table.index.name
    return df


def table_to_display(table: pd.DataFrame, periods: Iterable[str]) -> str:
    formatted = format_table(table)
    display_df = formatted.reset_index()
    index_name = formatted.index.name or "指标"
    if display_df.columns[0] != index_name:
        display_df = display_df.rename(columns={display_df.columns[0]: index_name})

    flattened_columns: list[str] = []
    for col in display_df.columns:
        if isinstance(col, tuple):
            flattened_columns.append("-".join(filter(None, map(str, col))))
        else:
            flattened_columns.append(str(col))
    display_df.columns = flattened_columns

    display_df, sep_columns = insert_period_separators(display_df, periods)
    style_columns = []
    for idx, col in enumerate(sep_columns, start=1):
        display_df = display_df.rename(columns={col: ""})
        style_col = f"{col}_style"
        display_df[style_col] = "bg-yellow"
        style_columns.append(style_col)

    display_df[index_name] = display_df[index_name].apply(
        lambda v: "" if v is None else str(v).strip()
    )

    def bold_data(value: Optional[object]) -> str:
        if value is None:
            return ""
        text = str(value).strip()
        if text == "" or text == "-" or text.startswith("【"):
            return text
        return f"**{text}**"

    for col in display_df.columns:
        if col == index_name or col in style_columns:
            continue
        display_df[col] = display_df[col].apply(bold_data)

    try:
        return display_df.drop(columns=style_columns).to_markdown(
            index=False,
            tablefmt="github",
            numalign="right",
            stralign="center",
        )
    except ImportError:
        return display_df.drop(columns=style_columns).to_string(
            index=False
        )


def format_table(table: pd.DataFrame) -> pd.DataFrame:
    formatted = table.copy().astype("object")
    for metric in formatted.index:
        for column in formatted.columns:
            formatted.at[metric, column] = format_value(metric, formatted.at[metric, column])
    formatted = add_section_headers(formatted)
    return formatted


def load_statement(stock_code: str, indicator: str) -> pd.DataFrame:
    df = ak.stock_financial_report_sina(stock=stock_code, symbol=indicator)
    df = df.copy()
    date_col = None
    for candidate in ("报表日期", "报告日", "报告期"):
        if candidate in df.columns:
            date_col = candidate
            break
    if date_col is None:
        raise KeyError("未在新浪财报数据中找到日期字段，请检查接口返回格式是否更新")
    df[date_col] = pd.to_datetime(df[date_col])
    df = df.set_index(date_col)
    df = df[~df.index.duplicated(keep="last")]
    return convert_numeric(df)


def fetch_datasets(symbol: str) -> Dataset:
    stock_code = normalize_stock(symbol)
    base_symbol = strip_exchange_prefix(stock_code)
    company_name, latest_price, total_shares, total_market_cap = fetch_company_overview(base_symbol)
    company_label = company_name or base_symbol.upper()
    latest_valuation = fetch_latest_valuation(base_symbol)
    return Dataset(
        symbol=symbol,
        stock_code=stock_code,
        company_name=company_label,
        profit=load_statement(stock_code, "利润表"),
        cashflow=load_statement(stock_code, "现金流量表"),
        balance=load_statement(stock_code, "资产负债表"),
        latest_price=latest_price,
        total_shares=total_shares,
        total_market_cap=total_market_cap,
        latest_valuation=latest_valuation,
    )


def safe_div(numerator: Optional[float], denominator: Optional[float]) -> Optional[float]:
    if numerator is None or denominator in (None, 0):
        return None
    try:
        return numerator / denominator
    except ZeroDivisionError:
        return None


def build_summary(dataset: Dataset, report_dates: Dict[str, str]) -> pd.DataFrame:
    profit = dataset.profit
    cashflow = dataset.cashflow
    balance = dataset.balance
    market_cap = dataset.total_market_cap
    period_market_caps = compute_period_market_caps(dataset, report_dates)
    latest_val = dataset.latest_valuation
    latest_pe = latest_val.pe if latest_val else None
    latest_pb = latest_val.pb if latest_val else None
    latest_ps = latest_val.ps if latest_val else None
    latest_time = latest_val.timestamp if latest_val else None
    current_price = latest_val.price if latest_val else None
    current_cap_value = latest_val.market_cap if latest_val else None

    if current_price is None and dataset.latest_price is not None:
        try:
            current_price = float(dataset.latest_price)
        except (TypeError, ValueError):
            current_price = None

    def format_amount(value: Optional[float]) -> Optional[str]:
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return None
        absolute = abs(value)
        if absolute >= 1e8:
            return f"{value / 1e8:.2f}亿"
        if absolute >= 1e4:
            return f"{value / 1e4:.2f}万"
        return f"{value:.2f}元"

    def format_time_value(value: Optional[object]) -> Optional[str]:
        if value is None:
            return None
        if isinstance(value, pd.Timestamp):
            value = value.to_pydatetime()
        if isinstance(value, datetime):
            if value.hour == 0 and value.minute == 0 and value.second == 0:
                return value.strftime("%Y-%m-%d")
            return value.strftime("%Y-%m-%d %H:%M:%S")
        try:
            ts = pd.Timestamp(value)
        except Exception:
            text = str(value).strip()
            return text or None
        return format_time_value(ts)

    fallback_context: Optional[dict[str, object]] = None
    latest_period_label = next(iter(report_dates)) if report_dates else None

    revenue_col = find_column(profit, ["营业总收入", "营业收入"])
    cost_col = find_column(profit, ["营业成本"])
    operating_profit_col = find_column(profit, ["营业利润"])
    net_profit_col = find_column(profit, ["净利润"])
    parent_net_col = find_column(
        profit,
        ["归属于母公司所有者的净利润", "归属于母公司股东的净利润"],
    )

    ocf_col = find_column(cashflow, ["经营活动产生的现金流量净额", "经营活动现金流量净额"])
    capex_col = find_column(cashflow, ["购建固定资产、无形资产和其他长期资产支付的现金"])
    invest_cf_col = find_column(cashflow, ["投资活动产生的现金流量净额"])
    cash_sales_col = find_column(cashflow, ["销售商品、提供劳务收到的现金"])

    asset_total_col = find_column(balance, ["资产总计", "资产合计"])
    debt_total_col = find_column(balance, ["负债合计", "负债总计"])
    current_assets_col = find_column(balance, ["流动资产合计"])
    current_liabilities_col = find_column(balance, ["流动负债合计"])
    equity_col = find_column(
        balance,
        ["归属于母公司股东权益合计", "归属于母公司所有者权益合计"],
    )

    rows: list[Dict[str, Optional[float]]] = []

    for label, date_str in report_dates.items():
        row: Dict[str, Optional[float]] = {"期间": label}

        income = get_value(profit, date_str, revenue_col, 0)
        income_prev = get_value(profit, date_str, revenue_col, 1)
        income_prev_prev = get_value(profit, date_str, revenue_col, 2)

        income_growth = calc_growth(income, income_prev)
        income_growth_prev = calc_growth(income_prev, income_prev_prev)

        cost = get_value(profit, date_str, cost_col, 0)
        cost_prev = get_value(profit, date_str, cost_col, 1)

        gross_profit = safe_sub(income, cost)
        gross_profit_prev = safe_sub(income_prev, cost_prev)
        gross_margin = safe_div(gross_profit, income)
        gross_margin_prev = safe_div(gross_profit_prev, income_prev)

        operating_profit = get_value(profit, date_str, operating_profit_col, 0)
        operating_profit_prev = get_value(profit, date_str, operating_profit_col, 1)

        net_profit = get_value(profit, date_str, net_profit_col, 0)
        net_profit_prev = get_value(profit, date_str, net_profit_col, 1)
        net_profit_growth = calc_growth(net_profit, net_profit_prev)

        parent_net = get_value(profit, date_str, parent_net_col, 0)
        parent_net_prev = get_value(profit, date_str, parent_net_col, 1)

        ocf = get_value(cashflow, date_str, ocf_col, 0)
        ocf_prev = get_value(cashflow, date_str, ocf_col, 1)

        capex_raw = get_value(cashflow, date_str, capex_col, 0)
        capex_prev_raw = get_value(cashflow, date_str, capex_col, 1)
        capex = -capex_raw if capex_raw is not None else None
        capex_prev = -capex_prev_raw if capex_prev_raw is not None else None

        invest_cf = get_value(cashflow, date_str, invest_cf_col, 0)
        invest_cf_prev = get_value(cashflow, date_str, invest_cf_col, 1)

        cash_sales = get_value(cashflow, date_str, cash_sales_col, 0)
        cash_sales_prev = get_value(cashflow, date_str, cash_sales_col, 1)

        fcf = ocf + capex_raw if ocf is not None and capex_raw is not None else None
        fcf_prev = (
            ocf_prev + capex_prev_raw
            if ocf_prev is not None and capex_prev_raw is not None
            else None
        )

        cash_coverage = safe_div(ocf, net_profit)
        cash_coverage_prev = safe_div(ocf_prev, net_profit_prev)

        income_cash_ratio = safe_div(cash_sales, income)
        income_cash_ratio_prev = safe_div(cash_sales_prev, income_prev)

        total_assets = get_value(balance, date_str, asset_total_col, 0)
        total_assets_prev = get_value(balance, date_str, asset_total_col, 1)
        total_assets_prev_prev = get_value(balance, date_str, asset_total_col, 2)

        total_debt = get_value(balance, date_str, debt_total_col, 0)
        total_debt_prev = get_value(balance, date_str, debt_total_col, 1)

        debt_ratio = safe_div(total_debt, total_assets)
        debt_ratio_prev = safe_div(total_debt_prev, total_assets_prev)

        current_assets = get_value(balance, date_str, current_assets_col, 0)
        current_assets_prev = get_value(balance, date_str, current_assets_col, 1)

        current_liabilities = get_value(balance, date_str, current_liabilities_col, 0)
        current_liabilities_prev = get_value(balance, date_str, current_liabilities_col, 1)

        current_ratio = safe_div(current_assets, current_liabilities)
        current_ratio_prev = safe_div(current_assets_prev, current_liabilities_prev)

        net_assets = get_value(balance, date_str, equity_col, 0)
        net_assets_prev = get_value(balance, date_str, equity_col, 1)
        net_assets_prev_prev = get_value(balance, date_str, equity_col, 2)

        avg_equity = mean_valid(net_assets, net_assets_prev)
        avg_equity_prev = mean_valid(net_assets_prev, net_assets_prev_prev)

        avg_assets = mean_valid(total_assets, total_assets_prev)
        avg_assets_prev = mean_valid(total_assets_prev, total_assets_prev_prev)
        total_asset_turnover = safe_div(income, avg_assets)
        total_asset_turnover_prev = safe_div(income_prev, avg_assets_prev)

        equity_multiplier = safe_div(avg_assets, avg_equity)
        equity_multiplier_prev = safe_div(avg_assets_prev, avg_equity_prev)

        net_profit_margin = safe_div(net_profit, income)
        net_profit_margin_prev = safe_div(net_profit_prev, income_prev)

        parent_net_margin = safe_div(parent_net, income)
        parent_net_margin_prev = safe_div(parent_net_prev, income_prev)

        roe = safe_product([parent_net_margin, total_asset_turnover, equity_multiplier])
        if roe is None:
            roe = safe_div(parent_net, avg_equity)
        roe_prev = safe_product(
            [parent_net_margin_prev, total_asset_turnover_prev, equity_multiplier_prev]
        )
        if roe_prev is None:
            roe_prev = safe_div(parent_net_prev, avg_equity_prev)

        report_ts = pd.Timestamp(date_str)
        prev_year_end = f"{report_ts.year - 1}-12-31"
        net_profit_prev_year_end = get_value_at(profit, prev_year_end, net_profit_col)
        ttm_net_profit = calc_ttm(net_profit, net_profit_prev, net_profit_prev_year_end)

        period_info = period_market_caps.get(label)
        period_cap_value: Optional[float] = None
        period_price_value: Optional[float] = None
        period_time_value: Optional[object] = None
        used_latest_price = False

        if period_info:
            period_cap_value = period_info.get("market_cap")
            period_price_value = period_info.get("price")
            period_time_value = period_info.get("price_time")
        else:
            period_cap_value = market_cap

        total_shares = dataset.total_shares
        if period_price_value is None and period_cap_value is not None and total_shares not in (None, 0):
            try:
                period_price_value = float(period_cap_value) / float(total_shares)
            except (TypeError, ValueError, ZeroDivisionError):
                period_price_value = None

        if (
            period_price_value is None
            and dataset.latest_price is not None
        ):
            try:
                period_price_value = float(dataset.latest_price)
                used_latest_price = True
            except (TypeError, ValueError):
                period_price_value = None

        if period_time_value is None:
            if used_latest_price and latest_time is not None:
                period_time_value = latest_time
            else:
                try:
                    period_time_value = pd.Timestamp(date_str)
                except Exception:
                    period_time_value = latest_time

        pe = safe_div(period_cap_value, ttm_net_profit) if period_cap_value is not None else None
        pb = safe_div(period_cap_value, net_assets) if period_cap_value is not None else None
        ps = safe_div(period_cap_value, income) if period_cap_value is not None else None

        row.update(
            {
                "营业收入": income,
                "营业收入-去年同期": income_prev,
                "营业收入增长率": income_growth,
                "营业收入增长率-去年同期": income_growth_prev,
                "毛利率": gross_margin,
                "毛利率-去年同期": gross_margin_prev,
                "营业利润率": safe_div(operating_profit, income),
                "营业利润率-去年同期": safe_div(operating_profit_prev, income_prev),
                "净利润": net_profit,
                "净利润-去年同期": net_profit_prev,
                "净利润率": net_profit_margin,
                "净利润率-去年同期": net_profit_margin_prev,
                "归母净利润": parent_net,
                "归母净利润-去年同期": parent_net_prev,
                "归母净利润率": parent_net_margin,
                "归母净利润率-去年同期": parent_net_margin_prev,
                "总资产周转率": total_asset_turnover,
                "总资产周转率-去年同期": total_asset_turnover_prev,
                "权益乘数": equity_multiplier,
                "权益乘数-去年同期": equity_multiplier_prev,
                "ROE": roe,
                "ROE-去年同期": roe_prev,
                "经营活动现金流量净额": ocf,
                "经营活动现金流量净额-去年同期": ocf_prev,
                "净利润现金保障倍数": cash_coverage,
                "净利润现金保障倍数-去年同期": cash_coverage_prev,
                "资本性支出": capex,
                "资本性支出-去年同期": capex_prev,
                "投资活动现金流量净额": invest_cf,
                "投资活动现金流量净额-去年同期": invest_cf_prev,
                "自由现金流": fcf,
                "自由现金流-去年同期": fcf_prev,
                "销售商品提供劳务收到的现金": cash_sales,
                "销售商品提供劳务收到的现金-去年同期": cash_sales_prev,
                "现金收入比率": income_cash_ratio,
                "现金收入比率-去年同期": income_cash_ratio_prev,
                "资产负债率": debt_ratio,
                "资产负债率-去年同期": debt_ratio_prev,
                "流动比率": current_ratio,
                "流动比率-去年同期": current_ratio_prev,
                "净资产": net_assets,
                "净资产-去年同期": net_assets_prev,
                "市值": period_cap_value,
                "PE": pe,
                "PB": pb,
                "PS": ps,
                "最新PE": latest_pe,
                "最新PB": latest_pb,
                "最新PS": latest_ps,
                "估值": None,
                "估值时间": period_time_value,
                "最新估值": None,
                "最新估值时间": None,
            }
        )

        if (
            fallback_context is None
            or pd.Timestamp(date_str) > fallback_context.get("ts", pd.Timestamp.min)
        ):
            fallback_context = {
                "ts": pd.Timestamp(date_str),
                "ttm_net_profit": ttm_net_profit,
                "net_assets": net_assets,
                "income": income,
            }

        cap_text = format_amount(period_cap_value)
        price_text_for_row = format_amount(period_price_value)
        if cap_text and price_text_for_row:
            row["市值"] = f"{cap_text}（股价：{price_text_for_row}）"
        elif cap_text:
            row["市值"] = cap_text
        elif price_text_for_row:
            row["市值"] = f"股价：{price_text_for_row}"
        else:
            row["市值"] = None

        summary_parts_for_period: list[str] = []
        if price_text_for_row:
            summary_parts_for_period.append(f"股价：{price_text_for_row}")
        if cap_text:
            summary_parts_for_period.append(f"市值：{cap_text}")
        if summary_parts_for_period:
            row["估值"] = " | ".join(summary_parts_for_period)
        else:
            row["估值"] = None

        rows.append(row)

    summary_df = pd.DataFrame(rows).set_index("期间")

    if "估值时间" in summary_df.columns:
        summary_df["估值时间"] = summary_df["估值时间"].apply(format_time_value)
    if "最新估值时间" in summary_df.columns:
        summary_df["最新估值时间"] = summary_df["最新估值时间"].apply(format_time_value)

    for col in ("市值", "估值", "估值时间", "最新估值", "最新估值时间"):
        if col in summary_df.columns:
            summary_df[col] = summary_df[col].astype("object")

    if summary_df.empty:
        return summary_df

    def is_missing(value: Optional[float]) -> bool:
        return value is None or (isinstance(value, float) and pd.isna(value))

    if fallback_context is not None:
        if current_cap_value is None:
            current_cap_value = dataset.total_market_cap
        if current_cap_value is None and dataset.latest_price is not None and dataset.total_shares not in (None, 0):
            try:
                current_cap_value = float(dataset.latest_price) * float(dataset.total_shares)
            except (TypeError, ValueError):
                current_cap_value = None

        if current_price is None and current_cap_value is not None and dataset.total_shares not in (None, 0):
            try:
                current_price = float(current_cap_value) / float(dataset.total_shares)
            except (TypeError, ValueError):
                current_price = None

        if current_cap_value is not None:
            fallback_pe = safe_div(current_cap_value, fallback_context.get("ttm_net_profit"))
            fallback_pb = safe_div(current_cap_value, fallback_context.get("net_assets"))
            fallback_ps = safe_div(current_cap_value, fallback_context.get("income"))

            if is_missing(latest_pe) and fallback_pe is not None:
                latest_pe = fallback_pe
                summary_df["最新PE"] = latest_pe
            if is_missing(latest_pb) and fallback_pb is not None:
                latest_pb = fallback_pb
                summary_df["最新PB"] = latest_pb
            if is_missing(latest_ps) and fallback_ps is not None:
                latest_ps = fallback_ps
                summary_df["最新PS"] = latest_ps

            if latest_time is None or str(latest_time).strip() == "":
                latest_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    latest_summary_text = None
    price_text = format_amount(current_price) if current_price is not None else None
    cap_text = format_amount(current_cap_value) if current_cap_value is not None else None
    time_text_formatted = format_time_value(latest_time)

    summary_parts = []
    if price_text:
        summary_parts.append(f"股价：{price_text}")
    if cap_text:
        summary_parts.append(f"市值：{cap_text}")
    if summary_parts:
        latest_summary_text = " | ".join(summary_parts)

    if latest_period_label and latest_period_label in summary_df.index:
        if latest_summary_text is not None:
            summary_df.at[latest_period_label, "最新估值"] = latest_summary_text
        if time_text_formatted is not None:
            summary_df.at[latest_period_label, "最新估值时间"] = time_text_formatted

    return summary_df


def display_dataframe(table: pd.DataFrame, periods: Iterable[str]) -> None:
    print("\n====== 财务指标汇总 ======")
    print(table_to_display(table, periods))
    print()


def maybe_export_excel(
    table: pd.DataFrame,
    export_path: Optional[Path],
    dataset: Dataset,
    periods: Iterable[str],
) -> None:
    if export_path is None:
        safe_company = sanitize_filename_component(dataset.company_name) or "company"
        base_symbol = strip_exchange_prefix(dataset.stock_code).upper()
        filename = f"{safe_company}_{base_symbol}.xlsx"
        export_path = Path.cwd() / filename
    export_path.parent.mkdir(parents=True, exist_ok=True)

    formatted_table = format_table(table)
    excel_table = formatted_table.reset_index()
    index_name = formatted_table.index.name or "指标"
    if excel_table.columns[0] != index_name:
        excel_table = excel_table.rename(columns={excel_table.columns[0]: index_name})

    flattened_columns: list[str] = []
    for col in excel_table.columns:
        if isinstance(col, tuple):
            flattened_columns.append("-".join(filter(None, map(str, col))))
        else:
            flattened_columns.append(str(col))
    excel_table.columns = flattened_columns

    excel_table, sep_columns = insert_period_separators(excel_table, periods)
    sep_display_names: dict[str, str] = {col: "" for idx, col in enumerate(sep_columns)}
    excel_table = excel_table.rename(columns=sep_display_names)
    separator_labels = set(sep_display_names.values())

    sheet_name = "财务指标"

    with pd.ExcelWriter(export_path) as writer:
        excel_table.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]

        ws.freeze_panes = "B2"
        ws.sheet_view.showGridLines = False
        ws.row_dimensions[1].height = 39

        header_font = Font(name="微软雅黑", color="333333")
        header_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_side = Side(style="thin", color="DDDDDD")
        header_border = Border(top=thin_side, bottom=thin_side, left=thin_side, right=thin_side)
        indicator_fill = PatternFill(start_color="F0F5FF", end_color="F0F5FF", fill_type="solid")
        body_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        indicator_font = Font(name="微软雅黑", color="204060")
        body_font = Font(name="微软雅黑", color="333333")
        section_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        section_font = Font(name="微软雅黑", color="7F6000")
        separator_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        zebra_fill = PatternFill(start_color="EDF2FA", end_color="EDF2FA", fill_type="solid")

        excel_columns = list(excel_table.columns)
        data_row_counter = 0

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = header_border

        for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            first_value = row_cells[0].value
            other_values = [cell.value for cell in row_cells[1:]]
            is_blank_row = all(
                (cell.value is None) or (str(cell.value).strip() == "") for cell in row_cells
            )
            is_section_row = (
                not is_blank_row
                and isinstance(first_value, str)
                and first_value.strip().startswith("【")
                and first_value.strip().endswith("】")
                and all((val is None) or (str(val).strip() == "") for val in other_values)
            )

            if is_blank_row:
                ws.row_dimensions[row_cells[0].row].height = 8
            elif is_section_row:
                ws.row_dimensions[row_cells[0].row].height = 39
            else:
                ws.row_dimensions[row_cells[0].row].height = 33
                data_row_counter += 1
            apply_zebra = (not is_blank_row and not is_section_row and data_row_counter % 2 == 1)

            for cell in row_cells:
                column_label = excel_columns[cell.col_idx - 1]
                if column_label in separator_labels:
                    cell.value = ""
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = Border()
                    cell.fill = separator_fill
                    continue
                if is_blank_row:
                    cell.value = ""
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = Border()
                    cell.fill = PatternFill(fill_type=None)
                    continue
                if is_section_row:
                    if cell.col_idx == 1:
                        cell.font = section_font
                        cell.fill = section_fill
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    else:
                        cell.value = ""
                        cell.fill = PatternFill(fill_type=None)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = Border()
                    continue
                alignment = Alignment(
                    horizontal="left" if cell.col_idx == 1 else "right",
                    vertical="center",
                )
                cell.alignment = alignment
                cell.border = body_border
                if cell.col_idx == 1:
                    cell.font = indicator_font
                    cell.fill = indicator_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.font = Font(name="微软雅黑", color="333333")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "0.00"
                    if apply_zebra:
                        cell.fill = zebra_fill

        for col_idx, column in enumerate(excel_columns, start=1):
            column_letter = get_column_letter(col_idx)
            if column in separator_labels:
                ws.column_dimensions[column_letter].width = 3
                continue
            max_length = calc_display_length(column)
            for value in excel_table[column].fillna(""):
                max_length = max(max_length, calc_display_length(value))
            if col_idx == 1:
                width = max(20, min(44, max_length + 6))
            else:
                width = max(18, min(40, max_length + 6))
            ws.column_dimensions[column_letter].width = width

    print(f"已导出 Excel：{export_path}")


def main() -> None:
    args = parse_args()
    report_dates = build_report_dates(args.report)
    ordered_periods = list(report_dates.keys())
    symbols: list[str] = []
    if args.symbols:
        symbols.extend(args.symbols)
    if args.symbol:
        symbols.append(args.symbol)
    if not symbols:
        symbols = ["000002"]

    excel_arg = Path(args.excel) if args.excel else None
    multi = len(symbols) > 1
    excel_dir: Optional[Path] = None

    if excel_arg:
        if multi:
            if excel_arg.suffix:
                excel_dir = excel_arg.parent / excel_arg.stem
            else:
                excel_dir = excel_arg
            excel_dir.mkdir(parents=True, exist_ok=True)
        else:
            if excel_arg.is_dir() or not excel_arg.suffix:
                excel_dir = excel_arg
                excel_dir.mkdir(parents=True, exist_ok=True)
            else:
                excel_dir = None

    for idx, symbol in enumerate(symbols, start=1):
        dataset = fetch_datasets(symbol)
        summary = build_summary(dataset, report_dates)
        combined_table = combine_summary(summary, ordered_periods)

        export_path: Optional[Path] = None
        if excel_arg:
            if multi:
                target_dir = excel_dir or excel_arg
                target_dir.mkdir(parents=True, exist_ok=True)
                export_path = target_dir / default_excel_filename(dataset)
            else:
                if excel_dir:
                    export_path = excel_dir / default_excel_filename(dataset)
                else:
                    export_path = excel_arg
        else:
            # 如果没有指定 Excel 路径，使用默认文件名
            export_path = Path.cwd() / default_excel_filename(dataset)

        maybe_export_excel(combined_table, export_path, dataset, ordered_periods)


if __name__ == "__main__":
    main()

